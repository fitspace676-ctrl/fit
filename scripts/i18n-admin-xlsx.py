#!/usr/bin/env python3
"""Export / import the `admin` i18n namespace as an XLSX for manual translation.

  export:  python3 scripts/i18n-admin-xlsx.py export [out.xlsx]
  import:  python3 scripts/i18n-admin-xlsx.py import <in.xlsx>

The sheet carries one row per leaf key (`admin.*`) with the English source, the
current Georgian value and the ICU placeholders that must survive translation.
Importing writes column D (or column C when D is empty) back into
`packages/i18n/locales/ka.json`, preserving key order and file formatting.
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
LOCALES = ROOT / "packages" / "i18n" / "locales"
EN = LOCALES / "en.json"
KA = LOCALES / "ka.json"
NAMESPACE = "admin"

# Matches the variable name that opens an ICU argument — `{name}`, `{name, plural, …}`.
# Deliberately ignores the branch text inside a plural/select, which is prose and
# is *supposed* to differ between locales.
PLACEHOLDER = re.compile(r"\{\s*([A-Za-z_][A-Za-z0-9_]*)\s*[,}]")


def flatten(node, prefix=""):
    for key, value in node.items():
        path = f"{prefix}.{key}" if prefix else key
        if isinstance(value, dict):
            yield from flatten(value, path)
        else:
            yield path, value


HEADERS = ["Key", "Section", "English (en)", "ქართული (ka)", "Placeholders"]

# The console chrome — nav, tab strips, the profile menu. Not a page of its own,
# so these live together on one sheet instead of six three-row ones.
SHELL = (
    "common",
    "nav",
    "navGroups",
    "profile",
    "system",
    "classesTabs",
    "paymentsTabs",
    "invoicesHub",
    "ptHub",
)
SHELL_SHEET = "00 · Shell & nav"

# Sheet titles per admin page. Anything not listed falls back to the raw
# namespace segment, so a newly added page still gets its own sheet.
PAGE_TITLES = {
    "dashboard": "Dashboard",
    "members": "Members",
    "staff": "Staff",
    "trainers": "Trainers",
    "schedule": "Schedule",
    "pos": "POS",
    "checkin": "Check-in",
    "marketing": "Marketing",
    "loyalty": "Loyalty",
    "automation": "Automation",
    "settings": "Settings",
    "locations": "Locations",
    "billingPlans": "Billing plans",
    "analytics": "Analytics",
    "reports": "Reports",
    "activity": "Activity log",
    "agent": "AI agent",
}


def style_sheet(ws) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2937")
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    for col, width in {"A": 52, "B": 26, "C": 60, "D": 60, "E": 22}.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"


def export(out_path: Path) -> None:
    en = json.loads(EN.read_text(encoding="utf-8"))
    ka = json.loads(KA.read_text(encoding="utf-8"))

    en_flat = dict(flatten(en[NAMESPACE], NAMESPACE))
    ka_flat = dict(flatten(ka.get(NAMESPACE, {}), NAMESPACE))

    # Bucket every key by admin page, keeping en.json's original key order.
    pages: dict[str, list[str]] = {}
    for key in en_flat:
        page = key.split(".")[1]
        pages.setdefault(SHELL_SHEET if page in SHELL else page, []).append(key)

    shell = pages.pop(SHELL_SHEET, [])
    ordered = [(SHELL_SHEET, shell)] if shell else []
    # Biggest pages first, so the bulk of the work is up front.
    ordered += sorted(pages.items(), key=lambda item: -len(item[1]))

    wrap = Alignment(wrap_text=True, vertical="top")
    mono = Font(name="Menlo", size=10)

    wb = Workbook()
    index = wb.active
    index.title = "Index"
    index.append(["Sheet", "Namespace", "Keys"])

    position = 0
    for page, keys in ordered:
        if page == SHELL_SHEET:
            title = page
        else:
            position += 1
            title = f"{position:02d} · {PAGE_TITLES.get(page, page)}"
        ws = wb.create_sheet(title[:31])
        ws.append(HEADERS)

        for key in keys:
            value = en_flat[key]
            # admin.classes.form.title -> "classes / form"
            parts = key.split(".")
            section = " / ".join(parts[1:-1]) if len(parts) > 2 else parts[1]
            names = dict.fromkeys(PLACEHOLDER.findall(str(value)))
            ws.append(
                [
                    key,
                    section,
                    value,
                    ka_flat.get(key, ""),
                    ", ".join(f"{{{name}}}" for name in names),
                ]
            )
            row = ws.max_row
            ws.cell(row=row, column=1).font = mono
            for col in (3, 4):
                ws.cell(row=row, column=col).alignment = wrap
            ws.cell(row=row, column=5).font = mono

        style_sheet(ws)
        index.append([ws.title, f"{NAMESPACE}.{page}" if page != SHELL_SHEET else "—", len(keys)])

    index.cell(row=1, column=1).font = Font(bold=True, color="FFFFFF")
    for col in range(1, 4):
        index.cell(row=1, column=col).font = Font(bold=True, color="FFFFFF")
        index.cell(row=1, column=col).fill = PatternFill("solid", fgColor="1F2937")
    index.append(["TOTAL", "", len(en_flat)])
    index.cell(row=index.max_row, column=1).font = Font(bold=True)
    index.cell(row=index.max_row, column=3).font = Font(bold=True)
    for col, width in {"A": 28, "B": 26, "C": 10}.items():
        index.column_dimensions[col].width = width
    index.freeze_panes = "A2"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    missing = sum(1 for k in en_flat if not ka_flat.get(k))
    print(
        f"wrote {out_path}  ({len(en_flat)} keys across {len(ordered)} sheets, "
        f"{missing} with no Georgian value)"
    )


def set_path(node: dict, key: str, value: str) -> None:
    parts = key.split(".")
    for part in parts[:-1]:
        node = node.setdefault(part, {})
    node[parts[-1]] = value


def do_import(in_path: Path) -> None:
    en = json.loads(EN.read_text(encoding="utf-8"))
    ka = json.loads(KA.read_text(encoding="utf-8"))
    en_flat = dict(flatten(en[NAMESPACE], NAMESPACE))

    wb = load_workbook(in_path, read_only=True)

    translated: dict[str, str] = {}
    unknown: list[str] = []
    for ws in wb.worksheets:
        # Skip the Index sheet (and anything else without the key/value layout).
        header = [cell.value for cell in next(ws.iter_rows(max_row=1), [])]
        if not header or header[0] != HEADERS[0] or len(header) < 4:
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            key = (row[0] or "").strip()
            if not key:
                continue
            value = row[3] if len(row) > 3 and row[3] not in (None, "") else row[2]
            if value in (None, ""):
                continue
            if key not in en_flat:
                unknown.append(key)
                continue
            translated[key] = str(value)

    # Rebuild the admin namespace in en.json's key order so the diff stays clean.
    rebuilt: dict = {}
    missing: list[str] = []
    ka_flat = dict(flatten(ka.get(NAMESPACE, {}), NAMESPACE))
    for key in en_flat:
        value = translated.get(key) or ka_flat.get(key)
        if value is None:
            missing.append(key)
            continue
        set_path(rebuilt, key, value)

    ka[NAMESPACE] = rebuilt[NAMESPACE]
    KA.write_text(json.dumps(ka, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print(f"updated {KA} ({len(translated)} values from the sheet)")
    if missing:
        print(f"  ! {len(missing)} keys had no value at all: {missing[:5]}")
    if unknown:
        print(f"  ! {len(unknown)} rows are not admin.* keys in en.json: {unknown[:5]}")

    # Cheap placeholder sanity check.
    broken = [
        k
        for k, v in translated.items()
        if set(PLACEHOLDER.findall(en_flat[k])) != set(PLACEHOLDER.findall(v))
    ]
    if broken:
        print(f"  ! {len(broken)} values changed their ICU placeholders: {broken[:10]}")


if __name__ == "__main__":
    mode = sys.argv[1] if len(sys.argv) > 1 else "export"
    if mode == "export":
        target = Path(sys.argv[2]) if len(sys.argv) > 2 else ROOT / "admin-i18n-en-ka.xlsx"
        export(target)
    elif mode == "import":
        do_import(Path(sys.argv[2]))
    else:
        sys.exit(f"unknown mode: {mode}")
